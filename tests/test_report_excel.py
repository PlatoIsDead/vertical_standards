"""№10: xlsx-матрица «сотрудник × курс» — чистый модуль, читаем из bytes."""
import json
from io import BytesIO

from openpyxl import load_workbook

import app.report_excel as rx

Q10 = json.dumps({"basic_questions": [], "exam_questions": [
    {"id": i, "text": "?", "options": ["A. 1", "B. 2", "C. 3", "D. 4"],
     "correct": "A"} for i in range(10)]})     # порог сдачи round(7.0)=7


def _course(cid, name):
    return {"id": cid, "doc_name": name, "questions_json": Q10}


def _row(rid, uid, cid, state, exam=0, name=None, role=None):
    return {"id": rid, "user_id": uid, "course_id": cid, "state": state,
            "score_basic": 3, "score_exam": exam,
            "doc_name": name or f"Курс{cid}.docx", "questions_json": Q10,
            "updated_at": "2026-08-11T10:00:00", "role": role}


def _load(data):
    return load_workbook(BytesIO(data))


def test_matrix_cells_best_of_pair():
    courses = [_course(1, "ALL Первый.docx"), _course(2, "ALL Второй.docx"),
               _course(3, "ALL Третий.docx"), _course(4, "ALL Четвёртый.docx")]
    rows = [
        _row(1, "10", 1, "DONE", exam=8, name="ALL Первый.docx"),   # сдан
        _row(2, "10", 2, "DONE", exam=2, name="ALL Второй.docx"),   # провал…
        _row(3, "10", 2, "DONE", exam=9, name="ALL Второй.docx"),   # …пересдал
        _row(4, "10", 3, "DONE", exam=1, name="ALL Третий.docx"),   # не сдан
        _row(5, "10", 4, "READING", name="ALL Четвёртый.docx"),     # в процессе
    ]
    emps = {"10": {"full_name": "Иван Иванов", "work_position": "Админ"}}
    ws = _load(rx.build_report_xlsx(rows, emps, courses))["Матрица"]
    header = [c.value for c in ws[1]]
    assert header[:3] == ["Сотрудник", "Должность", "Отдел/роль"]
    assert header[3:] == ["Первый", "Второй", "Третий", "Четвёртый"]
    line = [c.value for c in ws[2]]
    assert line[0] == "Иван Иванов" and line[1] == "Админ"
    assert line[3:] == ["✅", "✅", "❌", "🕐"]     # лучший результат пары


def test_archived_course_and_legacy_uid_present():
    rows = [_row(1, "10", 5, "DONE", exam=8, name="ALL Архивный.docx")]
    ws = _load(rx.build_report_xlsx(rows, {}, []))["Матрица"]  # активных нет
    assert [c.value for c in ws[1]][3] == "Архивный"           # колонка жива
    line = [c.value for c in ws[2]]
    assert line[0] == "ID 10" and line[3] == "✅"               # legacy-строка


def test_not_started_empty_and_details_sheet():
    courses = [_course(1, "ALL Первый.docx")]
    emps = {"20": {"full_name": "Пётр Петров", "work_position": None}}
    wb = _load(rx.build_report_xlsx([], emps, courses))
    assert [c.value for c in wb["Матрица"][2]][3] in ("", None)  # не начинал
    assert "Детали" in wb.sheetnames

    wb2 = _load(rx.build_report_xlsx(
        [_row(1, "20", 1, "DONE", exam=9, name="ALL Первый.docx")],
        emps, courses))
    det = wb2["Детали"]
    line = [c.value for c in det[2]]
    assert line[0] == "Пётр Петров" and "9/10" in line[4]
