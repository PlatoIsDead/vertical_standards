"""
app/report_excel.py — №10: xlsx-отчёт «сотрудник × курс» (матрица Marriott).

Чистый модуль: bytes на выходе, без сети и БД. Правило ячейки зафиксировано:
сдал ХОТЬ ОДНУ сессию пары → ✅; иначе судим ПОСЛЕДНЮЮ (max id): DONE → ❌,
не-DONE → 🕐; сессий нет → пусто. Формула сдачи — единственная в репо:
score_exam >= round(exam_total * 0.7).
"""
import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.hr_tools import _session_status
from app.roles import display_name, role_name


def _questions(obj: dict) -> dict:
    try:
        return json.loads(obj.get("questions_json") or "{}")
    except json.JSONDecodeError:
        return {}


def _passed(session: dict, questions: dict) -> bool:
    total = len(questions.get("exam_questions", []))
    return bool(total) and session["score_exam"] >= round(total * 0.7)


def _cell(pair_sessions: list[dict] | None, questions: dict) -> str:
    if not pair_sessions:
        return ""
    if any(s["state"] == "DONE" and _passed(s, questions)
           for s in pair_sessions):
        return "✅"                                   # лучший результат пары
    last = max(pair_sessions, key=lambda s: s["id"])
    return "❌" if last["state"] == "DONE" else "🕐"


def _autowidth(ws) -> None:
    for i in range(1, ws.max_column + 1):
        letter = get_column_letter(i)
        width = max((len(str(c.value)) for c in ws[letter] if c.value), default=8)
        ws.column_dimensions[letter].width = min(width + 2, 40)


def build_report_xlsx(rows: list[dict], employees_by_uid: dict[str, dict],
                      courses: list[dict]) -> bytes:
    """rows = get_report_rows(); courses = get_active_courses().

    Колонки = активные курсы ∪ курсы с сессиями (архивный с историей сдач
    не выпадает). Строки = employees ∪ legacy-uid из сессий."""
    course_by_id = {c["id"]: c for c in courses}
    for r in rows:
        course_by_id.setdefault(
            r["course_id"],
            {"id": r["course_id"], "doc_name": r["doc_name"],
             "questions_json": r.get("questions_json")})
    course_cols = sorted(course_by_id.values(), key=lambda c: c["id"])

    uids = sorted(employees_by_uid,
                  key=lambda u: employees_by_uid[u].get("full_name") or f"~{u}")
    for r in rows:
        uid = str(r["user_id"])
        if uid not in employees_by_uid and uid not in uids:
            uids.append(uid)                          # legacy-сессии без записи

    by_pair: dict[tuple, list] = {}
    last_role: dict[str, str] = {}
    for r in sorted(rows, key=lambda r: r["id"]):
        by_pair.setdefault((str(r["user_id"]), r["course_id"]), []).append(r)
        if r.get("role"):
            last_role[str(r["user_id"])] = r["role"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Матрица"
    ws.append(["Сотрудник", "Должность", "Отдел/роль"]
              + [display_name(c["doc_name"]) for c in course_cols])
    for uid in uids:
        emp = employees_by_uid.get(uid, {})
        role = last_role.get(uid)
        ws.append([emp.get("full_name") or f"ID {uid}",
                   emp.get("work_position") or "",
                   role_name(role) if role else ""]
                  + [_cell(by_pair.get((uid, c["id"])), _questions(c))
                     for c in course_cols])
    ws.freeze_panes = "B2"
    _autowidth(ws)

    det = wb.create_sheet("Детали")
    det.append(["Сотрудник", "Курс", "Статус", "Базовый", "Экзамен", "Дата"])
    for r in sorted(rows, key=lambda r: r.get("updated_at") or "",
                    reverse=True):
        emp = employees_by_uid.get(str(r["user_id"]), {})
        q = _questions(r)
        det.append([
            emp.get("full_name") or f"ID {r['user_id']}",
            display_name(r["doc_name"]),
            _session_status(r, q),
            f"{r['score_basic']}/{len(q.get('basic_questions', []))}",
            f"{r['score_exam']}/{len(q.get('exam_questions', []))}",
            (r.get("updated_at") or "")[:10],
        ])
    _autowidth(det)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
